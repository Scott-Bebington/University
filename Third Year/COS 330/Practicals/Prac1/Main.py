import warnings

warnings.filterwarnings('ignore', category=DeprecationWarning)

from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font


sentence = "be at the third pillar from the left outside the lyceum theatre tonight at seven. if you are distrustful bring two friends."

new_sentence = sentence.replace(". ", "xx")
new_sentence = new_sentence.replace(".", "xx")
new_sentence = new_sentence.replace(" ", "")
split_sentence = list(new_sentence)
starting_letter = "B"
starting_number = 6
wb = Workbook()
# grab the active worksheet
ws = wb.active

boldFont = Font(bold=True, italic=False, underline='single')

ws['B2'] = "Author"
ws['C2'] = 'Scott Bebington'
ws['B3'] = "Student Number"
ws['C3'] = 'u21546216'
ws['B4'] = "COS 330"
ws['C4'] = "Practical 1"

memoryWords = ["cryptographic", "network security"]
# 
for word in memoryWords:

    # Remove duplicate characters from memoryWord1
    memoryWord = "".join(dict.fromkeys(word))
    memoryWord = memoryWord.replace(" ", "")

    if len(memoryWord) > 10:
        memoryWord = memoryWord[:10]

    print(f"Memory word: {memoryWord}")

    columnOrder = []
    alphabet = "abcdefghijklmnopqrstuvwxyz"
    for letter in alphabet:
        if letter in memoryWord:
            columnOrder.append(memoryWord.index(letter))


    print(f"Column order: {columnOrder}")

    # Create a 2D array to store the letters of new_sentence
    array_2d = []
    row = []
    for letter in split_sentence:
        row.append(letter)
        if len(row) == 10:
            array_2d.append(row)
            row = []

    # Add the last row if it contains less than 10 letters
    if row:
        while len(row) < 10:
            row.append('x')
        array_2d.append(row)

    # Print the 2D array
    for row in array_2d:
        print(row)

    word_array = []
    word = ""

    for col in columnOrder:
        for row in array_2d:
            if col < len(row):
                word += row[col]
                if len(word) == 5:
                    word_array.append(word)
                    word = ""

    print()
    print(word_array)

    joined_word = ''.join(word_array)
    print(joined_word)
    print("\n")

    # Create a workbook and add a worksheet.
    

    

    border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )

    cell = ws[starting_letter + str(starting_number)]
    cell.alignment = cell.alignment.copy(
        horizontal='center', vertical='center')
    cell.border = border

    ws.merge_cells(starting_letter + str(starting_number) + ':' +
                   chr(ord(starting_letter) + 9) + str(starting_number + 1))
    ws[starting_letter + str(starting_number)].alignment = ws[starting_letter + str(
        starting_number)].alignment.copy(horizontal='center', vertical='center')
    
    if (starting_number == 6):
        ws[starting_letter + str(starting_number)] = "Single Transposition Cypher"
    else:
        ws[starting_letter + str(starting_number)] = "Double Transposition Cypher"

    starting_number += 3

    for i in range(0, 10):
        letterEquivalent = chr(ord(starting_letter) + columnOrder[i])
        ws[letterEquivalent + str(starting_number)] = i + 1

    # starting_number = 3
    starting_number += 1

    for i in range(0, len(memoryWord)):
        letter = memoryWord[i]
        wsLetter = chr(ord(starting_letter) + i)
        ws[wsLetter + str(starting_number)] = letter
        ws[wsLetter + str(starting_number)].font = boldFont
        ws[wsLetter + str(starting_number)].border = border

    # starting_number = 4
    starting_number += 1

    for i in range(0, len(split_sentence), 10):
        row = split_sentence[i:i+10]
        for j in range(0, min(len(row), 10)):
            ws[chr(ord(starting_letter) + j) + str(starting_number)] = row[j]
        # Fill the remaining cells in the row with 'x'
        for k in range(len(row), 10):
            ws[chr(ord(starting_letter) + k) + str(starting_number)] = 'x'
        starting_number += 1

    # starting_number = 16
    starting_number += 1

    words_per_line = 10
    for i in range(0, len(word_array)):
        row = word_array[i]
        column = i % words_per_line
        row_number = starting_number + i // words_per_line
        ws[chr(ord(starting_letter) + column) + str(row_number)] = row

    # starting_number = 20
    starting_number += 4

    ws[starting_letter + str(starting_number)] = joined_word

    ws.merge_cells(starting_letter + str(starting_number) + ':' +
                   chr(ord(starting_letter) + 9) + str(starting_number + 1))
    ws[starting_letter + str(starting_number)].alignment = ws[starting_letter + str(
        starting_number)].alignment.copy(wrapText=True, horizontal='left', vertical='center')

    
    starting_number += 3

    split_sentence = []
    for word in word_array:
        split_sentence.extend(list(word))
        
# Center the text for every cell
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = cell.alignment.copy(horizontal='center', vertical='center')

# Save the file
    wb.save("Cypher Transposition.xlsx")